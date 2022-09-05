import openpyxl, csv, sqlite3, re, gender_guesser.detector as gender, pyautogui as gui, time
from gett_gender import get_gender, dev_num, form, a_form, get_time

#tentar adicionar ao SQLite
path = "C:/Users/João/Desktop/Access/Devedor.xlsx"
wb = openpyxl.Workbook() 
wb_obj = openpyxl.load_workbook(path) 
sheet = wb_obj.active

#https://www.geeksforgeeks.org/how-to-delete-one-or-more-rows-in-excel-using-openpyxl/
#maybe get a program to delete uneeded info with excel openpyxl
row = sheet.max_row
column = sheet.max_column
  
print("Total Rows:", row)
print("Total Columns:", column)

#------------------------------------------------------------------------------------------------------------
#descide which method to store contact number
#by the dev_num() takes obs and forms

with open("C:/Users/João/Desktop/Access/Devedor.csv", "r", encoding="Latin-1") as file:  
            csv_reader = csv.reader(file, delimiter=';')  
            line_count = 0 
            for devedor in csv_reader:
                if line_count == 0:  
                    line_count += 1   
                else:
                    horario = get_time() 
                    nome = devedor[1]
                    prim_nome = nome.split(' ')[0]
                    cobrador = devedor[24]
                    obs = devedor[29]
                    #also check for cli. cod, e setor
                    if cobrador == "6":
                        pronomes_dev = get_gender(prim_nome)
                        print("--------------------------------------------------------")
                        formando = form(obs)
                        print(f"Nome:{nome}, cod cobrador:{cobrador}, formando: {formando}")
                        #NEED TO CLEAN  FORMANDO INSITE OBS, TAKE ANY BULSHIT
                        forms = a_form(formando)
                        if forms == True:
                            pronomes_form = get_gender(formando)
                            text =f"""{horario} {pronomes_form[0]}, é o Vitor, referente ao seu album de formatura em aberto com a Millenium no nome {pronomes_dev[1]} {pronomes_dev[0]}{nome}. Estamos com condição especial para quitação *SEM JUROS*.
                            Caso haja interesse em negociar para limpar o nome {pronomes_dev[1]} {prim_nome}, me retorne para que eu te passe as condições."""
                            print(f"THIS IS ITTTTTTTTTTTTTTT {text}")
                        elif forms == False:
                            text =f"""{horario} {pronomes_dev[0]} {nome}, é o Vitor, referente ao seu album de formatura em aberto com a Millenium. Estamos com condição especial para quitação *SEM JUROS*.
                            Caso haja interesse em negociar para limpar seu nome, me retorne para que eu te passe as condições."""
                            print(text)
                            
                    line_count += 1
            print(f"Total Rows: {line_count}")

conn = sqlite3.connect(':memory:')
c = conn.cursor()

c.execute("""CREATE TABLE devedores (
    nome text, 
    formando text,
    numero real
    ) """)

